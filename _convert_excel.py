# -*- coding: utf-8 -*-
"""Convert TECSA inventory xlsx into InventoryApp seed data."""
from __future__ import annotations

import json
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

XLSX = Path(r"c:\Users\Tecsa-user\Desktop\excel\Інвентаризація техніки ТОВ ТЕКСА.xlsx")
OUT_JSON = Path(r"C:\Users\Tecsa-user\Documents\InventoryApp\inventory-import.json")
OUT_SEED = Path(r"C:\Users\Tecsa-user\Documents\InventoryApp\seed-data.js")

GENERAL_SITE = "Загальний склад"
LOCAL_SITES = ["Ігорівська", "Автозаводська", "Ірпінь"]
SITES = [GENERAL_SITE, *LOCAL_SITES]

NOW = datetime.now(timezone.utc).isoformat()


def clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in {"nan", "none", "null", "?"}:
        return ""
    return s


def uid() -> str:
    return str(uuid.uuid4())


def map_site(raw: str, note_parts: list[str]) -> str:
    s = clean(raw)
    if not s:
        return GENERAL_SITE

    low = s.lower().replace("ʼ", "'").replace("’", "'")

    if "автозаводськ" in low and "ігор" in low:
        note_parts.append(f"Місце (як у файлі): {s}")
        return "Ігорівська"

    if "автозаводськ" in low:
        if "," in s or "каб" in low:
            note_parts.append(f"Уточнення місця: {s}")
        return "Автозаводська"

    if "ірпін" in low:
        if "," in s or "каб" in low:
            note_parts.append(f"Уточнення місця: {s}")
        return "Ірпінь"

    if "ігор" in low:  # Ігорівська / Ігорвіська / Ігорівька
        if "," in s or "каб" in low or s != "Ігорівська":
            note_parts.append(f"Уточнення місця: {s}")
        return "Ігорівська"

    note_parts.append(f"Місце (як у файлі): {s}")
    return GENERAL_SITE


def is_broken(tech_state: str, capability: str) -> bool:
    blob = f"{tech_state} {capability}".lower()
    if not blob.strip():
        return False
    # explicit not working
    if re.search(r"\bне\s*робоч", blob) or re.search(r"\bне\s*працю", blob):
        return True
    if "не працює" in blob or "неробоч" in blob:
        return True
    return False


def is_working_flag(tech_state: str, capability: str) -> bool:
    if is_broken(tech_state, capability):
        return False
    # partial still counts as "справний з проблемами"
    return True


def extract_problems(tech_state: str, capability: str, found: str, other: str) -> str:
    parts = []
    for p in (found, other):
        p = clean(p)
        if p:
            parts.append(p)

    cap = clean(capability)
    low = cap.lower()
    if cap and (
        "частково" in low
        or "не працює" in low
        or "проблем" in low
        or "застар" in low
        or "потрібн" in low
        or "—" in cap
        or "-" in cap and "працює" in low
    ):
        # avoid duplicating plain "працює"
        if low not in {"працює", "робочій стан"}:
            parts.append(cap)

    # dedupe preserve order
    seen = set()
    out = []
    for p in parts:
        key = p.lower()
        if key not in seen:
            seen.add(key)
            out.append(p)
    return "\n".join(out)


def extract_action(useful: str, category: str, recommended: str) -> str:
    parts = []
    for p in (recommended, useful, category):
        p = clean(p)
        if p:
            parts.append(p)
    seen = set()
    out = []
    for p in parts:
        key = p.lower()
        if key not in seen:
            seen.add(key)
            out.append(p)
    return "\n".join(out)


def extract_note(presence: str, remarks: str, serial: str, extra: list[str]) -> str:
    parts = []
    if clean(serial):
        parts.append(f"№ п/п у Excel: {clean(serial)}")
    if clean(presence):
        parts.append(f"Фактична наявність: {clean(presence)}")
    if clean(remarks):
        parts.append(clean(remarks))
    parts.extend(extra)
    return "\n".join(parts)


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Інвентаризація"]

    employees_by_name: dict[str, str] = {}
    items = []

    for r in range(2, ws.max_row + 1):
        name = clean(ws.cell(r, 2).value)
        if not name:
            continue

        serial = clean(ws.cell(r, 1).value)
        inv = clean(ws.cell(r, 3).value)
        presence = clean(ws.cell(r, 4).value)
        tech_state = clean(ws.cell(r, 5).value)
        pib = clean(ws.cell(r, 6).value)
        place = clean(ws.cell(r, 7).value)
        remarks = clean(ws.cell(r, 8).value)
        capability = clean(ws.cell(r, 9).value)
        specs = clean(ws.cell(r, 10).value)
        useful = clean(ws.cell(r, 11).value)
        category = clean(ws.cell(r, 12).value)
        recommended = clean(ws.cell(r, 13).value)
        found = clean(ws.cell(r, 14).value)
        other = clean(ws.cell(r, 15).value)

        note_extra: list[str] = []
        site = map_site(place, note_extra)

        employee_id = ""
        if pib:
            if pib not in employees_by_name:
                employees_by_name[pib] = uid()
            employee_id = employees_by_name[pib]
        else:
            # no employee => on warehouse of mapped site (empty site => general)
            pass

        # model from specs if not identical to name
        model = specs if specs and specs.lower() != name.lower() else ""
        if not model and "(" in name:
            # keep model empty; name already contains model text
            model = ""

        problems = extract_problems(tech_state, capability, found, other)
        action = extract_action(useful, category, recommended)
        note = extract_note(presence, remarks, serial, note_extra)

        if tech_state:
            note_extra2 = []
            # keep tech state in note if useful and not already covered
            if tech_state.lower() not in {"робочий", "працює", "робочій"}:
                if tech_state not in note:
                    note = (note + "\n" if note else "") + f"Технічний стан: {tech_state}"

        item = {
            "id": uid(),
            "name": name,
            "model": model,
            "invNumber": inv,
            "working": is_working_flag(tech_state, capability),
            "site": site,
            "employeeId": employee_id,
            "specs": specs if specs else "",
            "note": note.strip(),
            "action": action,
            "problems": problems,
            "createdAt": NOW,
            "updatedAt": NOW,
            "source": "Інвентаризація",
        }
        items.append(item)

    # Sheet "Без С1"
    ws2 = wb["Без С1"]
    for r in range(1, ws2.max_row + 1):
        name = clean(ws2.cell(r, 1).value)
        if not name:
            continue
        status = clean(ws2.cell(r, 2).value)
        detail = clean(ws2.cell(r, 3).value)
        working = not is_broken(status, status)
        problems = ""
        if status and status.lower() not in {"працює"}:
            problems = status
        if detail:
            problems = (problems + "\n" if problems else "") + detail

        items.append(
            {
                "id": uid(),
                "name": name,
                "model": "",
                "invNumber": "",
                "working": working,
                "site": GENERAL_SITE,
                "employeeId": "",
                "specs": "",
                "note": "Джерело: аркуш «Без С1»",
                "action": "",
                "problems": problems.strip(),
                "createdAt": NOW,
                "updatedAt": NOW,
                "source": "Без С1",
            }
        )

    employees = [{"id": eid, "name": name} for name, eid in sorted(employees_by_name.items(), key=lambda x: x[0].lower())]

    # strip helper field
    for it in items:
        it.pop("source", None)

    payload = {
        "version": "excel-tecsa-2026-07-15",
        "employees": employees,
        "items": items,
    }

    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    seed_js = (
        "// Auto-generated from Excel inventory — do not edit by hand\n"
        "window.INVENTORY_SEED = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\n"
    )
    OUT_SEED.write_text(seed_js, encoding="utf-8")

    warehouse = sum(1 for i in items if not i["employeeId"])
    broken = sum(1 for i in items if not i["working"])
    by_site = {}
    for i in items:
        by_site[i["site"]] = by_site.get(i["site"], 0) + 1

    summary = {
        "employees": len(employees),
        "items": len(items),
        "warehouse": warehouse,
        "broken": broken,
        "by_site": by_site,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
